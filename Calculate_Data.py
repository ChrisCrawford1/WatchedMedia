from Converter import Converter
import openpyxl
import plotly.graph_objects as go

class CalculateData:
    media = []

    def get_workbook(self):
        convert = Converter()
        return convert.from_csv_to_workbook()

    def analyse_workbook(self):
        workbook = openpyxl.load_workbook(self.get_workbook())
        active_sheet = workbook.worksheets[0]
        print('{} has {} columns and {} rows'.format(active_sheet, active_sheet.max_column, active_sheet.max_row) + '\n')

        for i in range(2, active_sheet.max_row):
            value = str(active_sheet.cell(column=1, row=i).value)
            value = value.split(':')

            if len(value) > 1:
                media_type = {
                    'type': 'tv_show',
                    'name': value[0]
                }
                if media_type not in self.media:
                    self.media.append(media_type)
            else:
                media_type = {
                    'type': 'movie',
                    'name': value[0]
                }
                self.media.append(media_type)

        self.get_breakdown()

    def get_breakdown(self):
        tv_count = 0
        movie_count = 0
        for item in self.media:
            if item['type'] == 'tv_show':
                tv_count += 1
            else:
                movie_count += 1

        stats = {
            'tv_count' : tv_count,
            'movie_count' : movie_count,
            'tv_percentage' : round((tv_count / len(self.media)) * 100, 1),
            'movie_percentage' : round((movie_count / len(self.media)) * 100, 1),
            'total_media' : len(self.media)
        }

        output_type = input('What output do you want? [Image, Raw]: ').lower()
        self.provide_output(stats, output_type.lower())

    def provide_output(self, stats, type):
        if type == 'raw':
             print('TV Shows Watched: {}({}%), Movies Watched:{}({}%)\nTotal Media Consumed: {}'
             .format(stats['tv_count'], stats['tv_percentage'], 
             stats['movie_count'], stats['movie_percentage'], stats['total_media']))

        if type == 'image':
            labels = ['TV Shows Watched','Movies Watched']
            values = [stats['tv_count'], stats['movie_count']]

            fig = go.Figure(data=[go.Pie(labels=labels, values=values)])
            fig.show()
